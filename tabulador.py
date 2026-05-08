"""
tabulador.py  —  Motor de tabulação universal  —  IFec RJ
==========================================================
Replica a lógica do TabIFec (R) em Python.

Comportamento para perguntas com "Outro. Qual?" codificado:
  - Opções fechadas + Outro ficam numa tabela só
  - "Outro" aparece em fonte normal com total de quem marcou
  - Categorias codificadas aparecem recuadas, em itálico, abaixo de Outro
"""

from __future__ import annotations
import re
import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings("ignore")


# ─────────────────────────────────────────────────────────────────────────────
# TIPOS
# ─────────────────────────────────────────────────────────────────────────────

TIPOS_LABEL = {
    "RU":     "Resposta Única",
    "RM":     "Resposta Múltipla",
    "ABERTA": "Aberta / Codificada",
    "MEDIA":  "Numérica (Média)",
    "NPS":    "NPS (0-10)",
    "IGNORAR":"Ignorar",
}


# ─────────────────────────────────────────────────────────────────────────────
# LEITURA E set_header
# ─────────────────────────────────────────────────────────────────────────────

def set_header(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Replica set_header() do TabIFec: combina linha 0 + linha 1."""
    row0, row1 = df_raw.iloc[0], df_raw.iloc[1]
    filled, last = [], ""
    for v in row0:
        if pd.notna(v) and str(v).strip() not in ("", "nan"):
            last = str(v).strip()
        filled.append(last)

    cols = []
    for q, r in zip(filled, row1):
        r_str = str(r).strip() if pd.notna(r) else ""
        combined = q + r_str
        combined = re.sub(r"\s*(Response$|Open-Ended Response$|NA$)\s*$", "", combined)
        combined = re.sub(r"\s*\[.*?\]", "", combined)
        cols.append(combined.strip())

    df = df_raw.iloc[2:].copy().reset_index(drop=True)
    seen = {}
    unique_cols = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            unique_cols.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            unique_cols.append(c)
    df.columns = unique_cols
    return df


def carregar_base(caminho: str) -> pd.DataFrame:
    xl = pd.ExcelFile(caminho)
    df_raw = pd.read_excel(caminho, header=None, sheet_name=xl.sheet_names[0])
    return set_header(df_raw)


# ─────────────────────────────────────────────────────────────────────────────
# PADRÕES
# ─────────────────────────────────────────────────────────────────────────────

_IGNORAR = {
    "respondent_id","collector_id","date_created","date_modified",
    "ip_address","email_address","first_name","last_name","custom_1",
    "Pesquisador:","Pesquisador","Entrevistador:","Entrevistador",
    "Comentário:","Comentário","Observações:","Observações",
    "Setor:","Setor","TOTAL","NPS","FX_IDADE",
}

_PAD_OUTRO_COL = re.compile(r"\?.*Outr[oa]|Outr[oa].*(\?|:|\.)\s*$", re.IGNORECASE)
_PAD_NPS       = re.compile(r"de 0 a 10|numa escala de 0|probabilidade.*recomendar", re.IGNORECASE)
_PAD_MEDIA     = re.compile(r"de 1 a \d|de \d a \d|quantas? (anos?|vezes?|pessoas?)|qual.*idade|renda.*r\$", re.IGNORECASE)
_PAD_OUTRO_VAL = re.compile(r"Outr[oa].*[:?.]", re.IGNORECASE)


def _e_outro(col: str) -> bool:
    return bool(_PAD_OUTRO_COL.search(col))


def _e_cod(col: str) -> bool:
    return col.endswith("_cod")


# ─────────────────────────────────────────────────────────────────────────────
# DETECÇÃO DE PERGUNTAS
# ─────────────────────────────────────────────────────────────────────────────

def _prefixo_pergunta(col: str) -> str:
    """
    Extrai o prefixo que identifica a pergunta-raiz de uma coluna.

    Estratégia em ordem de prioridade:
      1. Se houver '?' no nome → prefixo é tudo até (e incluindo) o '?'
         Ex: "Qual curso?Cabeleireiro" → "Qual curso?"
             "Qual curso? (19/10)"    → "Qual curso?"
      2. Se houver ' (' no nome → prefixo é tudo até o ' ('
         Ex: "Última palestra (19/10)" → "Última palestra"
      3. Caso contrário → a própria coluna é o prefixo (pergunta autônoma)
    """
    # Prioridade 1 — corta no '?'
    idx_q = col.find('?')
    if idx_q != -1:
        return col[:idx_q + 1]

    # Prioridade 2 — corta no ' ('
    idx_p = col.find(' (')
    if idx_p != -1:
        return col[:idx_p]

    return col


def detectar_perguntas(df: pd.DataFrame) -> list[dict]:
    """
    Agrupa colunas por pergunta-raiz de forma genérica.

    Regra principal: colunas que compartilham o mesmo prefixo
    (até o '?' ou ' (') são agrupadas como uma única pergunta RM.

    Exemplos:
      "Qual curso?Cabeleireiro"  ┐
      "Qual curso?Maquiador"     ├─ mesma raiz "Qual curso?"  → RM
      "Qual curso?Barbeiro"      ┘

      "Última palestra (19/10)"  ┐
      "Última palestra (20/10)"  ├─ mesma raiz "Última palestra" → RM
      "Última palestra (21/10)"  ┘
    """
    todas = list(df.columns)

    # Colunas candidatas (sem sistema, sem _cod, sem Original_)
    princ = [c for c in todas
             if c not in _IGNORAR
             and not c.startswith("Original_")
             and not _e_cod(c)]

    # ── Passo 1: mapear cada coluna ao seu prefixo-raiz ───────────────────────
    # prefixo_map: coluna → prefixo
    # grupos_prefixo: prefixo → [colunas]
    prefixo_map: dict[str, str] = {}
    grupos_prefixo: dict[str, list[str]] = {}

    for col in princ:
        pref = _prefixo_pergunta(col)
        prefixo_map[col] = pref
        grupos_prefixo.setdefault(pref, []).append(col)

    # ── Passo 2: dentro de cada grupo de prefixo, separar cols_princ / cols_outros
    # e resolver raiz_map para o startswith original (filhas de filhas)
    # Mantemos também o startswith original para colunas que são filhas diretas
    # (ex: "P06. Avalie os itens abaixo?Atendimento" onde há sub-colunas)
    raiz_map: dict[str, str] = {}
    for col in princ:
        pref = prefixo_map[col]
        raiz_map[col] = pref  # todas as colunas do grupo apontam para o prefixo

    # ── Passo 3: montar perguntas na ordem em que aparecem no DataFrame ────────
    perguntas, num = [], 1
    vistas_prefixos: set[str] = set()

    for col in todas:
        if col not in prefixo_map:
            continue
        pref = prefixo_map[col]
        if pref in vistas_prefixos:
            continue
        if col in _IGNORAR or col.startswith("Original_"):
            continue
        vistas_prefixos.add(pref)

        grupo = grupos_prefixo[pref]

        # Separa colunas "Outro" das principais
        cols_outros = [c for c in grupo if _e_outro(c)]
        cols_princ  = [c for c in grupo if c not in cols_outros]

        if not cols_princ:
            continue

        # Procurar coluna _cod associada
        col_cod = None
        for c in cols_outros:
            cand = c + "_cod"
            if cand in df.columns:
                col_cod = cand
                break
        if col_cod is None:
            pref_curto = pref[:25]
            for c in todas:
                if _e_cod(c) and c[:-4].startswith(pref_curto):
                    col_cod = c
                    break

        # Usa o prefixo como nome da pergunta (sem o sufixo da opção)
        # Se só há uma coluna no grupo e o prefixo == coluna, usa a coluna
        nome_pergunta = pref if len(grupo) > 1 else cols_princ[0]

        tipo = _detectar_tipo(df, nome_pergunta, cols_princ)
        if tipo == "IGNORAR":
            continue

        # Forçar RM quando o grupo tem mais de uma coluna principal
        if tipo not in ("NPS", "ABERTA") and len(cols_princ) > 1:
            tipo = "RM"

        # Nota de "Não soube avaliar"
        nota_nsa = ""
        if tipo in ("RU", "RM"):
            n_nsa = 0
            for cp in cols_princ:
                serie_cp = _col(df, cp).astype(str).str.strip()
                n_nsa += int(serie_cp.str.lower().str.contains(
                    r"não soube avaliar|nao soube avaliar", regex=True).sum())
            if n_nsa > 0:
                nota_nsa = (f'{n_nsa} pessoa{"(s)" if n_nsa > 1 else ""} '
                            f'respondeu/responderam "Não soube avaliar".')

        nota = _nota_padrao(tipo)
        if nota_nsa:
            nota = (nota + "\n" + nota_nsa) if nota else nota_nsa

        perguntas.append({
            "num":         f"P{num:02d}",
            "pergunta":    nome_pergunta,
            "tipo":        tipo,
            "colunas":     cols_princ,
            "cols_outros": cols_outros,
            "col_cod":     col_cod,
            "nota":        nota,
            "ativo":       True,
        })
        num += 1

    return perguntas


def _detectar_tipo(df: pd.DataFrame, raiz: str, cols: list[str]) -> str:
    raiz_l = raiz.lower()
    if raiz in _IGNORAR or raiz.startswith("Original_") or not raiz.strip():
        return "IGNORAR"
    if _PAD_NPS.search(raiz_l):
        return "NPS"

    cols_s = [c for c in cols if not _e_outro(c)]

    if _PAD_MEDIA.search(raiz_l) and cols_s:
        sample = pd.to_numeric(_col(df, cols_s[0]), errors="coerce").dropna()
        if len(sample) > 0 and sample.mean() > 0:
            return "MEDIA"

    if len(cols_s) > 1:
        if sum(1 for c in cols_s if _col(df, c).notna().any()) > 1:
            return "RM"

    if not cols_s:
        return "IGNORAR"

    serie = _col(df, cols_s[0]).dropna().astype(str).pipe(lambda s: s[s.str.strip() != ""])
    if len(serie) == 0:
        return "IGNORAR"
    if serie.nunique() / len(serie) > 0.35:
        return "ABERTA"
    return "RU"


def _nota_padrao(tipo: str) -> str:
    return {
        "RU":     "Pergunta com resposta única.",
        "RM":     "Pergunta com resposta múltipla.",
        "ABERTA": "Pergunta aberta (espontânea).",
        "MEDIA":  "Média das respostas.",
        "NPS":    "Índice de Recomendação (NPS).",
    }.get(tipo, "")


def _col(df: pd.DataFrame, nome: str) -> pd.Series:
    """Retorna sempre uma pd.Series, mesmo que a coluna seja duplicada."""
    resultado = df[nome]
    if isinstance(resultado, pd.DataFrame):
        return resultado.iloc[:, 0]
    return resultado


def _cols(df: pd.DataFrame, nomes: list) -> pd.DataFrame:
    """Retorna DataFrame com as colunas pedidas, desduplicando cada uma."""
    partes = [_col(df, n).rename(n) for n in nomes]
    return pd.concat(partes, axis=1)


# ─────────────────────────────────────────────────────────────────────────────
# TABULAÇÃO
# Retorna DataFrame com colunas: [" ", "Total", "%", "is_sub"]
# is_sub=True → linha em itálico recuado no Excel (sub-item de Outro)
# ─────────────────────────────────────────────────────────────────────────────

def tabular_ru_rm(df: pd.DataFrame, pergunta: dict) -> pd.DataFrame:
    cols_p   = pergunta["colunas"]
    cols_o   = pergunta.get("cols_outros", [])
    col_cod  = pergunta.get("col_cod")

    # ── Opções fechadas ───────────────────────────────────────────────────────
    df_long = (_cols(df, cols_p).stack().reset_index(drop=True).to_frame(name=" "))
    df_long[" "] = df_long[" "].astype(str).str.strip()
    df_long = df_long[
        df_long[" "].notna() &
        ~df_long[" "].isin(["", "-", "nan"]) &
        ~df_long[" "].str.startswith("NÃO SE APLICA") &
        ~df_long[" "].str.match(_PAD_OUTRO_VAL) &
        ~df_long[" "].str.lower().str.contains(r"não soube avaliar|nao soube avaliar", regex=True)
    ]

    freq = df_long[" "].value_counts().reset_index()
    freq.columns = [" ", "Total"]
    freq["is_sub"] = False

    # ── Contagem de Outro ─────────────────────────────────────────────────────
    n_outro = 0
    if cols_o:
        n_outro = int(
            _cols(df, cols_o).apply(
                lambda col: col.notna() & ~col.astype(str).str.strip().isin(["", "-", "nan"]),
                axis=0
            ).any(axis=1).sum()
        )
    elif col_cod and col_cod in df.columns:
        serie_outro = _col(df, col_cod)
        n_outro = int(
            (serie_outro.notna() &
             ~serie_outro.astype(str).str.strip().isin(["", "-", "nan"])).sum()
        )

    # ── Sub-itens codificados ─────────────────────────────────────────────────
    sub_df = pd.DataFrame(columns=[" ", "Total", "is_sub"])
    if col_cod and col_cod in df.columns:
        serie_cod = (
            _col(df, col_cod).dropna().astype(str)
            .pipe(lambda s: s[~s.str.strip().isin(["", "-", "nan"])])
        )
        if len(serie_cod) > 0:
            if not cols_o:
                n_outro = len(serie_cod)
            expandido = serie_cod.str.split(", ").explode().str.strip()
            expandido = expandido[expandido != ""]
            sf = expandido.value_counts().reset_index()
            sf.columns = [" ", "Total"]
            sf["is_sub"] = True
            sub_df = sf

    if n_outro > 0 or len(sub_df) > 0:
        outro = pd.DataFrame({" ": ["Outro"], "Total": [n_outro], "is_sub": [False]})
        freq  = pd.concat([freq, outro, sub_df], ignore_index=True)

    # ── Base (respondentes únicos) ────────────────────────────────────────────
    todas_cols = cols_p + cols_o
    respondentes = int(_cols(df, todas_cols).apply(
        lambda row: any(
            pd.notna(v) and str(v).strip() not in ("", "-", "NÃO SE APLICA")
            for v in row
        ), axis=1
    ).sum()) or 1

    total = pd.DataFrame({" ": ["Total"], "Total": [respondentes], "is_sub": [False]})
    freq  = pd.concat([freq, total], ignore_index=True)

    freq["%"] = freq.apply(
        lambda r: (r["Total"] / respondentes)
        if isinstance(r["Total"], (int, np.integer)) and r[" "] != "Total"
        else (1.0 if r[" "] == "Total" else "-"),
        axis=1
    )
    return freq[[" ", "Total", "%", "is_sub"]]


def tabular_aberta(df: pd.DataFrame, pergunta: dict, sep: str = ", ") -> pd.DataFrame:
    col_cod = pergunta.get("col_cod")
    col     = col_cod if (col_cod and col_cod in df.columns) else pergunta["colunas"][0]

    serie = (_col(df, col).dropna().astype(str)
             .pipe(lambda s: s[~s.str.strip().isin(["", "-", "nan"])]))

    expandido = serie.str.split(sep).explode().str.strip()
    expandido = expandido[expandido != ""]
    expandido = expandido[~expandido.str.match(_PAD_OUTRO_VAL)]

    freq = expandido.value_counts().reset_index()
    freq.columns = [" ", "Total"]
    freq["is_sub"] = False

    respondentes = len(serie) or 1
    total = pd.DataFrame({" ": ["Total"], "Total": [respondentes], "is_sub": [False]})
    freq  = pd.concat([freq, total], ignore_index=True)
    freq["%"] = freq["Total"].apply(
        lambda x: x / respondentes if isinstance(x, (int, np.integer)) else "-"
    )
    return freq[[" ", "Total", "%", "is_sub"]]


def tabular_media(df: pd.DataFrame, pergunta: dict) -> pd.DataFrame:
    valores = pd.to_numeric(_col(df, pergunta["colunas"][0]), errors="coerce").dropna()
    media   = round(float(valores.mean()), 2) if len(valores) > 0 else 0
    rows = [
        [" ", "Total", "%",  "is_sub"],
        ["Média",  media,        "-", False],
        ["Total",  len(valores), "-", False],
    ]
    return pd.DataFrame(rows[1:], columns=rows[0])


def tabular_nps(df: pd.DataFrame, pergunta: dict) -> pd.DataFrame:
    """
    Retorna 4 blocos separados por linha vazia, como no padrão:
      1. Distribuição de notas (0-10)
      2. Grupos (Promotores / Neutros / Detratores)
      3. NPS
      4. Média
    Cada bloco tem seu próprio cabeçalho e linha Total.
    Usamos is_sub=False para tudo; o exportar_excel os escreve em sequência.
    Sinalizamos a separação com linhas sentinela is_sep=True.
    """
    valores = pd.to_numeric(_col(df, pergunta["colunas"][0]), errors="coerce").dropna()
    total   = len(valores)
    if total == 0:
        return pd.DataFrame(columns=[" ", "Total", "%", "is_sub", "is_sep"])

    prom = int((valores >= 9).sum())
    neut = int(((valores >= 7) & (valores <= 8)).sum())
    detr = int((valores <= 6).sum())
    nps  = round((prom - detr) / total * 100, 1)
    media = round(float(valores.mean()), 6)

    rows = []

    # Bloco 1 — distribuição de notas (ordem decrescente)
    for nota_val, cnt in valores.value_counts().sort_index(ascending=False).items():
        rows.append((int(nota_val), int(cnt), int(cnt)/total, False, False))
    rows.append(("Total", total, 1.0, False, False))

    # Separador
    rows.append((None, None, None, False, True))

    # Bloco 2 — grupos
    rows.append(("Promotores",  prom,  prom/total,  False, False))
    rows.append(("Neutros",     neut,  neut/total,  False, False))
    rows.append(("Detratores",  detr,  detr/total,  False, False))
    rows.append(("Total",       total, 1.0,         False, False))

    # Separador
    rows.append((None, None, None, False, True))

    # Bloco 3 — NPS
    rows.append((f"NPS", round(nps, 5), "-", False, False))

    # Separador
    rows.append((None, None, None, False, True))

    # Bloco 4 — Média
    rows.append(("Média", media, "-", False, False))
    rows.append(("Total", total, "-", False, False))

    return pd.DataFrame(rows, columns=[" ", "Total", "%", "is_sub", "is_sep"])


def tabular_pergunta(df: pd.DataFrame, pergunta: dict) -> pd.DataFrame:
    t = pergunta["tipo"]
    if t in ("RU", "RM"): return tabular_ru_rm(df, pergunta)
    if t == "ABERTA":     return tabular_aberta(df, pergunta)
    if t == "MEDIA":      return tabular_media(df, pergunta)
    if t == "NPS":        return tabular_nps(df, pergunta)
    return pd.DataFrame(columns=[" ", "Total", "%", "is_sub"])


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAÇÃO EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def exportar_excel(df: pd.DataFrame, perguntas: list[dict],
                   saida: str, titulo: str = "Pesquisa",
                   total_respostas: int = None,
                   metodologia: str = None, rodape: str = None):

    from openpyxl.styles import Border, Side

    if total_respostas is None:
        total_respostas = len(df)

    _MET = metodologia or (
        "Metodologia: Pesquisa do tipo quantitativa com coleta de dados "
        "através de questionário (com perguntas abertas e fechadas) "
        "aplicado de forma presencial."
    )
    _ROD = rodape or (
        'Todas as perguntas foram obrigatórias. Nas perguntas com múltiplas respostas, '
        'os respondentes podem apontar mais de uma opção, por isso a soma das frequências '
        'passa de 100%. Nas perguntas com campo aberto na opção "Outro", as frequências '
        'apresentadas são das respostas válidas categorizadas. Os respondentes também podem '
        'apresentar mais de uma resposta no campo aberto "Outro".'
    )

    # ── Paleta do padrão Professional Fair ───────────────────────────────────
    COR_TITULO   = "984806"   # laranja escuro — título principal
    COR_DATA     = "E36C0A"   # laranja médio — data/subtítulo
    COR_PERGUNTA = "984806"   # laranja escuro — número/título da pergunta
    COR_TEXTO    = "262626"   # cinza muito escuro — dados
    COR_NOTA     = "404040"   # cinza médio — nota de rodapé
    COR_HDR_BG   = "E36C0A"   # fundo cabeçalho de tabela
    COR_HDR_FG   = "FFFFFF"   # texto cabeçalho de tabela

    FILL_HDR = PatternFill("solid", fgColor=COR_HDR_BG)
    AL_WRAP  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AL_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AL_IND   = Alignment(horizontal="left",   vertical="center", wrap_text=True, indent=2)

    BD_DOUBLE = Border(bottom=Side(border_style="double"))
    BD_MEDIUM = Border(top=Side(border_style="medium"),
                       bottom=Side(border_style="medium"))

    def F(bold=False, size=12, color=COR_TEXTO, italic=False, name="Cambria"):
        return Font(name=name, bold=bold, size=size, color=color, italic=italic)

    def num(v):
        if isinstance(v, (pd.Series, pd.DataFrame)):
            v = v.iloc[0] if len(v) > 0 else 0
        try:
            fv = float(v)
            return int(fv) if fv == int(fv) else round(fv, 2)
        except Exception:
            return v

    wb = Workbook()
    ws = wb.active
    ws.title = "Tabulação"

    row = 1

    # L1 — Título principal
    c = ws.cell(row=row, column=1, value=titulo)
    c.font = F(bold=True, size=18, color=COR_TITULO, name="Palatino")
    ws.row_dimensions[row].height = 22.5
    row += 1

    # L2 — Subtítulo/data
    subtitulo_txt = f"Total de respostas obtidas: {total_respostas}"
    c = ws.cell(row=row, column=1, value=subtitulo_txt)
    c.font = F(bold=True, size=12, color=COR_DATA, name="Cambria")
    ws.row_dimensions[row].height = 15.75
    row += 1  # L3 vazia
    ws.row_dimensions[row].height = 15.75
    row += 1

    # L4 — Metodologia
    c = ws.cell(row=row, column=1, value=_MET)
    c.font = F(bold=True, size=12, color=COR_TITULO, name="Palatino")
    c.alignment = AL_WRAP
    ws.row_dimensions[row].height = 15.75
    row += 1

    # L5 — Rodapé metodológico
    c = ws.cell(row=row, column=1, value=_ROD)
    c.font = F(bold=False, size=10, color=COR_NOTA, italic=True, name="Calibri")
    c.alignment = AL_WRAP
    ws.row_dimensions[row].height = 15.75
    row += 1  # L6 vazia
    ws.row_dimensions[row].height = 15.75
    row += 1

    # ── Perguntas ─────────────────────────────────────────────────────────────
    ativas = [p for p in perguntas if p.get("ativo", True) and p["tipo"] != "IGNORAR"]
    for p in ativas:
        tabela = tabular_pergunta(df, p)
        if tabela.empty:
            continue

        tipo = p["tipo"]

        # Título da pergunta (itálico, laranja, Calibri 12)
        c = ws.cell(row=row, column=1, value=f"{p['num']}. {p['pergunta']}")
        c.font = F(bold=False, size=12, color=COR_PERGUNTA, italic=True, name="Calibri")
        c.alignment = AL_WRAP
        ws.row_dimensions[row].height = 15.75
        row += 1

        # Cabeçalho de colunas com borda double embaixo
        hdrs = ["", "Total", "%"]
        extra_hdrs = []  # para NPS com colunas extras

        # Verifica se a tabela tem colunas de Detratores/Neutros/Promotores (vindo de NPS especial)
        # No padrão, P03 tem colunas extras E/F/G/H/I/J
        # Por ora mantemos as 3 colunas padrão
        for ci, hdr in enumerate(hdrs, 1):
            c = ws.cell(row=row, column=ci, value=hdr if hdr else None)
            c.font = F(bold=True, size=12, color=COR_HDR_FG, name="Cambria")
            c.fill = FILL_HDR
            c.alignment = AL_CTR
            c.border = BD_DOUBLE
            if hdr == "%":
                c.number_format = "0.0%"
        ws.row_dimensions[row].height = 16.5
        row += 1

        # Linhas de dados
        for _, dr in tabela.iterrows():
            # Separador entre blocos (ex: NPS com 4 sub-tabelas)
            if dr.get("is_sep", False):
                # Linha em branco + novo cabeçalho de colunas
                ws.row_dimensions[row].height = 8
                row += 1
                for ci, hdr in enumerate(["", "Total", "%"], 1):
                    c = ws.cell(row=row, column=ci, value=hdr if hdr else None)
                    c.font = F(bold=True, size=12, color=COR_HDR_FG, name="Cambria")
                    c.fill = FILL_HDR
                    c.alignment = AL_CTR
                    c.border = BD_DOUBLE
                    if hdr == "%":
                        c.number_format = "0.0%"
                ws.row_dimensions[row].height = 16.5
                row += 1
                continue

            label  = str(dr[" "]) if pd.notna(dr[" "]) else ""
            is_tot = label.strip().lower() == "total"
            is_sub = bool(dr.get("is_sub", False))
            is_nps = label in ("NPS",) or label.startswith("NPS =")

            # Fonte e alinhamento por tipo de linha
            if is_sub:
                f_lbl = F(italic=True, size=9, color="555555", name="Calibri")
                f_num = F(italic=True, size=9, color="555555", name="Calibri")
                a_lbl = AL_IND
                h = 13
            elif is_tot:
                f_lbl = F(bold=False, size=12, color=COR_TEXTO, name="Cambria")
                f_num = F(bold=False, size=12, color=COR_TEXTO, name="Cambria")
                a_lbl = AL_LEFT
                h = 16.5
            else:
                f_lbl = F(bold=False, size=12, color=COR_TEXTO, name="Cambria")
                f_num = F(bold=False, size=12, color=COR_TEXTO, name="Cambria")
                a_lbl = AL_LEFT
                h = 15.75

            # Coluna A — label
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = f_lbl
            c1.alignment = a_lbl
            if is_tot:
                c1.border = BD_MEDIUM

            # Coluna B — Total
            tv = dr["Total"]
            val_b = num(tv) if not is_nps and tv != "-" else tv
            c2 = ws.cell(row=row, column=2, value=val_b)
            c2.font = f_num
            c2.alignment = AL_CTR
            if is_tot:
                c2.border = BD_MEDIUM

            # Coluna C — %
            pv = dr["%"]
            if isinstance(pv, float) and not np.isnan(pv):
                val_c = pv
                c3_fmt = "0.0%"
            else:
                val_c = "-" if (is_tot and tipo == "RM") else str(pv) if pv is not None else "-"
                c3_fmt = "General"
            c3 = ws.cell(row=row, column=3, value=val_c)
            c3.font = f_num
            c3.alignment = AL_CTR
            c3.number_format = c3_fmt
            if is_tot:
                c3.border = BD_MEDIUM

            ws.row_dimensions[row].height = h
            row += 1

        # Nota de rodapé da pergunta
        if p.get("nota"):
            c = ws.cell(row=row, column=1, value=p["nota"])
            c.font = F(bold=False, size=10, color=COR_NOTA, italic=True, name="Calibri")
            c.alignment = AL_WRAP
            ws.row_dimensions[row].height = 15.75
            row += 1

        row += 1  # linha em branco entre perguntas

    # ── Dimensões das colunas (padrão) ────────────────────────────────────────
    ws.column_dimensions["A"].width = 90.71
    ws.column_dimensions["B"].width = 10.0
    ws.column_dimensions["C"].width = 9.14

    # ── Aba Base — cola o DataFrame bruto ─────────────────────────────────────
    ws_base = wb.create_sheet("Base")
    # Cabeçalho
    for ci, col_name in enumerate(df.columns, 1):
        c = ws_base.cell(row=1, column=ci, value=str(col_name))
        c.font = Font(name="Calibri", bold=True, size=10)
    # Dados
    for ri, (_, row_data) in enumerate(df.iterrows(), 2):
        for ci, val in enumerate(row_data, 1):
            ws_base.cell(row=ri, column=ci, value=val)

    wb.save(saida)
    return saida